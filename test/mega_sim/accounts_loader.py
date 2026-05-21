from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

from accounts.account_import import import_user_follows_from_sheet
from accounts.models import PlatformAccount, RegulatorAccount, UserAccount, WriterAccount


@dataclass(frozen=True)
class LoadedCounts:
    writers: int
    users: int
    platforms: int
    regulators: int


def _repo_root() -> Path:
    # .../课程沙盘/test/mega_sim/accounts_loader.py -> parents[2] == .../课程沙盘
    return Path(__file__).resolve().parents[2]


def load_accounts_from_excel(*, path: Optional[Path] = None) -> LoadedCounts:
    """
    从仓库内 `账号管理.xlsx` 加载账号到 pytest 测试库。
    注意：pytest 下 settings.BASE_DIR 会被 conftest 改为 tmp_path，因此这里不使用 BASE_DIR。
    """
    xlsx = Path(path) if path else (_repo_root() / "账号管理.xlsx")
    if not xlsx.exists():
        # 若仓库没有 Excel，则保留 conftest 的 pytest_* 账号作为 fallback
        return LoadedCounts(
            writers=WriterAccount.objects.count(),
            users=UserAccount.objects.count(),
            platforms=PlatformAccount.objects.count(),
            regulators=RegulatorAccount.objects.count(),
        )

    try:
        import openpyxl  # type: ignore
    except Exception:
        # 运行环境未安装 openpyxl：退化为使用既有账号，不阻塞大型自动化
        return LoadedCounts(
            writers=WriterAccount.objects.count(),
            users=UserAccount.objects.count(),
            platforms=PlatformAccount.objects.count(),
            regulators=RegulatorAccount.objects.count(),
        )

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws_writers = wb.worksheets[0]
        ws_users = wb.worksheets[1]
        ws_platforms = wb["平台"] if "平台" in wb.sheetnames else None
        ws_regulators = wb["监管"] if "监管" in wb.sheetnames else None

        def _iter_rows(ws):
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            # 跳过表头
            return rows[1:]

        w_cnt = u_cnt = p_cnt = r_cnt = 0

        for row in _iter_rows(ws_writers):
            if not row:
                continue
            account = str(row[0]).strip() if row[0] is not None else ""
            pwd = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not account or not pwd:
                continue
            WriterAccount.objects.update_or_create(账号=account, defaults={"密码": pwd})
            w_cnt += 1

        header_u = [
            str(x).strip() if x is not None else ""
            for x in (next(ws_users.iter_rows(min_row=1, max_row=1, values_only=True)) or [])
        ]
        col_u = {name: idx for idx, name in enumerate(header_u) if name}
        idx_u_acc = col_u.get("账号", 0)
        idx_u_pwd = col_u.get("密码", 1)
        idx_u_plat = col_u.get("所属平台")

        for row in _iter_rows(ws_users):
            if not row:
                continue
            account = str(row[idx_u_acc]).strip() if row[idx_u_acc] is not None else ""
            pwd = (
                str(row[idx_u_pwd]).strip()
                if idx_u_pwd is not None and idx_u_pwd < len(row) and row[idx_u_pwd] is not None
                else ""
            )
            if not account or not pwd:
                continue
            defaults = {"密码": pwd}
            if (
                idx_u_plat is not None
                and idx_u_plat < len(row)
                and row[idx_u_plat] is not None
            ):
                try:
                    defaults["所属平台"] = int(row[idx_u_plat])
                except (TypeError, ValueError):
                    pass
            UserAccount.objects.update_or_create(账号=account, defaults=defaults)
            u_cnt += 1

        writer_platform_by_account = dict(
            WriterAccount.objects.values_list("账号", "所属平台")
        )
        import_user_follows_from_sheet(
            ws_users,
            writer_platform_by_account=writer_platform_by_account,
        )

        if ws_platforms is not None:
            for row in _iter_rows(ws_platforms):
                if not row:
                    continue
                account = str(row[0]).strip() if row[0] is not None else ""
                pwd = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                platform_no = row[2] if len(row) > 2 else None
                if not account or not pwd:
                    continue
                defaults = {"密码": pwd}
                if platform_no is not None:
                    try:
                        defaults["所属平台"] = int(platform_no)
                    except Exception:
                        pass
                PlatformAccount.objects.update_or_create(账号=account, defaults=defaults)
                p_cnt += 1

        if ws_regulators is not None:
            for row in _iter_rows(ws_regulators):
                if not row:
                    continue
                account = str(row[0]).strip() if row[0] is not None else ""
                pwd = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if not account or not pwd:
                    continue
                RegulatorAccount.objects.update_or_create(账号=account, defaults={"密码": pwd})
                r_cnt += 1

        # 若 Excel 未提供平台/监管账号，则保留已有账号
        return LoadedCounts(
            writers=max(w_cnt, WriterAccount.objects.count()),
            users=max(u_cnt, UserAccount.objects.count()),
            platforms=PlatformAccount.objects.count(),
            regulators=RegulatorAccount.objects.count(),
        )
    finally:
        wb.close()

