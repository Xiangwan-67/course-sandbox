# -*- coding: utf-8 -*-
from io import BytesIO

import openpyxl
import pytest
from django.core.management import call_command

from accounts.account_import import (
    apply_user_initial_follows,
    import_user_follows_from_sheet,
    parse_follow_writer_accounts,
    sync_follow_fan_counts,
    validate_follow_same_platform,
)
from accounts.models import UserAccount, UserFollowWriter, WriterAccount


@pytest.mark.parametrize(
    'cell,expected',
    [
        (None, []),
        ('', []),
        ("['writer1', 'writer2']", ['writer1', 'writer2']),
        ('writer3, writer4', ['writer3', 'writer4']),
    ],
)
def test_parse_follow_writer_accounts(cell, expected):
    assert parse_follow_writer_accounts(cell) == expected


@pytest.mark.django_db
def test_validate_follow_same_platform_rejects_cross_platform():
    err = validate_follow_same_platform(
        user_account='user1',
        user_platform=0,
        writer_account='writer2',
        writer_platform_by_account={'writer2': 1},
    )
    assert err is not None
    assert '不能关注平台' in err


@pytest.mark.django_db
def test_apply_user_initial_follows_creates_and_counts():
    WriterAccount.objects.create(账号='w1', 密码='p', 所属平台=0, 粉丝数=0)
    WriterAccount.objects.create(账号='w2', 密码='p', 所属平台=0, 粉丝数=0)
    user = UserAccount.objects.create(账号='u1', 密码='p', 所属平台=0, 关注数=0)
    plat_map = {'w1': 0, 'w2': 0}
    apply_user_initial_follows(user, ['w1', 'w2'], plat_map)
    sync_follow_fan_counts(writer_accounts=['w1', 'w2'], user_accounts=['u1'])
    assert UserFollowWriter.objects.filter(用户=user).count() == 2
    user.refresh_from_db()
    assert user.关注数 == 2
    w1 = WriterAccount.objects.get(账号='w1')
    assert w1.粉丝数 == 1


@pytest.mark.django_db
def test_load_accounts_imports_follows_from_excel(tmp_path):
    path = tmp_path / 'accounts.xlsx'
    wb = openpyxl.Workbook()
    ws_w = wb.active
    ws_w.title = '写手'
    ws_w.append(['账号', '密码', '所属平台'])
    ws_w.append(['writer_a', 'pw', 0])
    ws_w.append(['writer_b', 'pw', 0])
    ws_u = wb.create_sheet('用户')
    ws_u.append(['账号', '密码', '所属平台', '关注'])
    ws_u.append(['user_a', 'pw', 0, "['writer_a']"])
    ws_u.append(['user_b', 'pw', 0, "['writer_a', 'writer_b']"])
    wb.save(path)

    call_command('load_accounts', '--file', str(path), '--clear')
    assert UserFollowWriter.objects.count() == 3
    u1 = UserAccount.objects.get(账号='user_a')
    assert u1.关注数 == 1
    assert set(u1.关注列表.values_list('写手账号', flat=True)) == {'writer_a'}


@pytest.mark.django_db
def test_load_accounts_rejects_cross_platform_follow(tmp_path):
    path = tmp_path / 'bad.xlsx'
    wb = openpyxl.Workbook()
    ws_w = wb.active
    ws_w.append(['账号', '密码', '所属平台'])
    ws_w.append(['writer_x', 'pw', 1])
    ws_u = wb.create_sheet('用户')
    ws_u.append(['账号', '密码', '所属平台', '关注'])
    ws_u.append(['user_x', 'pw', 0, "['writer_x']"])
    wb.save(path)

    from io import StringIO

    err = StringIO()
    call_command('load_accounts', '--file', str(path), '--clear', stderr=err)
    assert '不能关注平台' in err.getvalue() or '平台' in err.getvalue()
    assert UserFollowWriter.objects.count() == 0


@pytest.mark.django_db
def test_import_user_follows_from_sheet_value_error():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['账号', '密码', '所属平台', '关注'])
    ws.append(['ghost_user', 'pw', 0, "['writer1']"])
    with pytest.raises(ValueError, match='未导入'):
        import_user_follows_from_sheet(ws, writer_platform_by_account={'writer1': 0})
