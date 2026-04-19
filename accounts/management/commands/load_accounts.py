# -*- coding: utf-8 -*-
"""
从 账号管理.xlsx 导入账号到「写手」「用户」「平台账号」「监管机构账号」表。

- Sheet1 -> 写手（列：账号、密码）
- Sheet2 -> 用户（列：账号、密码）
- Sheet「平台」（若存在）-> 平台账号（列：账号、密码、对应编号）
- Sheet「监管机构」（若存在）-> 监管机构（列：账号、密码、负责平台）
"""
import json
import os
from django.core.management.base import BaseCommand
from django.conf import settings
import openpyxl

from accounts.models import WriterAccount, UserAccount, PlatformAccount, RegulatorAccount
from accounts.platform_scope import validate_regulator_platform_list


class Command(BaseCommand):
    help = '从 账号管理.xlsx 导入写手/用户/平台账号（可重复运行覆盖更新）'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default=os.path.join(settings.BASE_DIR, '账号管理.xlsx'),
            help='Excel 文件路径',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='导入前清空现有写手与用户数据',
        )
        parser.add_argument(
            '--clear-platform',
            action='store_true',
            help='导入前清空现有平台账号数据',
        )
        parser.add_argument(
            '--clear-regulator',
            action='store_true',
            help='导入前清空现有监管机构账号数据',
        )

    def handle(self, *args, **options):
        path = options['file']
        if not os.path.isfile(path):
            self.stderr.write(self.style.ERROR(f'文件不存在: {path}'))
            return

        if options['clear']:
            WriterAccount.objects.all().delete()
            UserAccount.objects.all().delete()
            self.stdout.write('已清空写手、用户表。')

        if options.get('clear_platform'):
            PlatformAccount.objects.all().delete()
            self.stdout.write('已清空平台账号表。')

        if options.get('clear_regulator'):
            RegulatorAccount.objects.all().delete()
            self.stdout.write('已清空监管机构账号表。')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if len(wb.worksheets) < 2:
            self.stderr.write(self.style.ERROR('Excel 至少需要 2 个工作表。'))
            wb.close()
            return

        # Sheet1 -> 写手
        ws1 = wb.worksheets[0]
        rows1 = list(ws1.iter_rows(min_row=2, values_only=True))
        count_w = 0
        for row in rows1:
            if not row or len(row) < 2:
                continue
            account, password = str(row[0]).strip(), str(row[1] or '').strip()
            if not account:
                continue
            WriterAccount.objects.update_or_create(
                账号=account,
                defaults={'密码': password},
            )
            count_w += 1
        self.stdout.write(self.style.SUCCESS(f'写手: 导入 {count_w} 条。'))

        # Sheet2 -> 用户
        ws2 = wb.worksheets[1]
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        count_u = 0
        for row in rows2:
            if not row or len(row) < 2:
                continue
            account, password = str(row[0]).strip(), str(row[1] or '').strip()
            if not account:
                continue
            UserAccount.objects.update_or_create(
                账号=account,
                defaults={'密码': password},
            )
            count_u += 1
        self.stdout.write(self.style.SUCCESS(f'用户: 导入 {count_u} 条。'))

        # Sheet「平台」 -> 平台账号（若存在）
        ws_platform = wb['平台'] if '平台' in wb.sheetnames else (wb.worksheets[2] if len(wb.worksheets) >= 3 else None)
        if ws_platform is None:
            self.stdout.write('平台账号: 未找到 Sheet「平台」，已跳过。')
        else:
            # 表头：账号、密码、对应编号（平台编号）
            header = [str(x).strip() if x is not None else '' for x in (next(ws_platform.iter_rows(min_row=1, max_row=1, values_only=True)) or [])]
            col_map = {name: idx for idx, name in enumerate(header) if name}
            idx_account = col_map.get('账号', 0)
            idx_password = col_map.get('密码', 1)
            idx_platform = col_map.get('对应编号', 2)
            rows_p = list(ws_platform.iter_rows(min_row=2, values_only=True))
            count_p = 0
            for row in rows_p:
                if not row:
                    continue
                account = str(row[idx_account] or '').strip() if idx_account < len(row) else ''
                password = str(row[idx_password] or '').strip() if idx_password < len(row) else ''
                if not account:
                    continue
                try:
                    platform_id = int(row[idx_platform]) if idx_platform < len(row) and row[idx_platform] is not None else 0
                except (TypeError, ValueError):
                    platform_id = 0
                PlatformAccount.objects.update_or_create(
                    账号=account,
                    defaults={'密码': password, '所属平台': platform_id},
                )
                count_p += 1
            self.stdout.write(self.style.SUCCESS(f'平台账号: 导入 {count_p} 条。'))

        # Sheet「监管机构」-> 监管机构账号（若存在）
        if '监管机构' in wb.sheetnames:
            ws_reg = wb['监管机构']
            header_r = [str(x).strip() if x is not None else '' for x in (next(ws_reg.iter_rows(min_row=1, max_row=1, values_only=True)) or [])]
            cmap = {name: idx for idx, name in enumerate(header_r) if name}
            ia = cmap.get('账号', 0)
            ipw = cmap.get('密码', 1)
            ipl = cmap.get('负责平台', 2)
            rows_r = list(ws_reg.iter_rows(min_row=2, values_only=True))
            parsed = []
            for row in rows_r:
                if not row:
                    continue
                acc = str(row[ia] or '').strip() if ia < len(row) else ''
                pwd = str(row[ipw] or '').strip() if ipw < len(row) else ''
                cell_pl = row[ipl] if ipl < len(row) else None
                if not acc:
                    continue
                pl = self._parse_regulator_platform_list(cell_pl)
                parsed.append((acc, pwd, pl))

            seen_platforms = set()
            for acc, pwd, pl in parsed:
                for x in pl:
                    if x in seen_platforms:
                        self.stderr.write(
                            self.style.ERROR(f'监管机构 Excel：平台编号 {x} 在表内被多个机构重复负责')
                        )
                        wb.close()
                        return
                    seen_platforms.add(x)

            count_r = 0
            for acc, pwd, pl in parsed:
                existing = RegulatorAccount.objects.filter(账号=acc).first()
                ex_pk = existing.pk if existing else None
                err = validate_regulator_platform_list(pl, exclude_pk=ex_pk)
                if err:
                    self.stderr.write(self.style.ERROR(f'监管机构「{acc}」: {err}'))
                    wb.close()
                    return
                RegulatorAccount.objects.update_or_create(
                    账号=acc,
                    defaults={'密码': pwd, '负责平台编号列表': pl},
                )
                count_r += 1
            self.stdout.write(self.style.SUCCESS(f'监管机构: 导入 {count_r} 条。'))
        else:
            self.stdout.write('监管机构: 未找到 Sheet「监管机构」，已跳过。')

        wb.close()

    @staticmethod
    def _parse_regulator_platform_list(cell):
        if cell is None:
            return []
        s = str(cell).strip()
        if not s:
            return []
        try:
            v = json.loads(s)
            if isinstance(v, list):
                return sorted({int(x) for x in v})
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
        s2 = s.strip('[]').replace('，', ',')
        parts = [p.strip() for p in s2.split(',') if p.strip()]
        out = set()
        for p in parts:
            try:
                out.add(int(p))
            except (TypeError, ValueError):
                continue
        return sorted(out)
